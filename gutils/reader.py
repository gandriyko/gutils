from django.utils.translation import gettext as _
from django.utils.encoding import force_str
from openpyxl.reader.excel import load_workbook
from django.utils.functional import cached_property
from gutils import Struct
import mmap
import dbf
import xlrd
import re
import csv
import os
import copy
from _csv import Error as CSVError


R_0 = re.compile(r'\.0$')
MAX_ERRORS = 30


def detect_encoding(value, force_cp1251=False):
    dos = list(range(0x80, 0xB0)) + list(range(0xE0, 0xF0))
    win = list(range(0xC0, 0x100))
    win_count = dos_count = 0
    if not force_cp1251:
        try:
            value.decode('utf8')
            return 'utf8'
        except Exception:
            pass
    for char in value:
        if not isinstance(char, int):
            char = ord(char)
        if char in dos:
            dos_count += 1
        if char in win:
            win_count += 1

    if win_count >= dos_count:
        return 'cp1251'
    else:
        return 'cp866'


def detect_delimiter(raw_value):
    delimiters = {b';': 0, b'\t': 0, b',': 0, b':': 0}
    for char in delimiters.keys():
        delimiters[char] = raw_value.count(char)
    res = sorted(delimiters, key=delimiters.get, reverse=True)
    return force_str(res[0])


def detect_quoting(raw_value):
    # max lines 500
    if raw_value.count(b'"') >= 200:
        return csv.QUOTE_MINIMAL
    return csv.QUOTE_NONE


def count_lines(filename):
    if not os.path.getsize(filename):
        return 0
    f = open(filename, "r+")
    buf = mmap.mmap(f.fileno(), 0)
    lines = 0
    readline = buf.readline
    while readline():
        lines += 1
    return lines


class ExcelReader(object):

    def __init__(self, filename, **kwargs):
        if not os.path.isfile(filename):
            raise NameError("%s is not a valid filename" % filename)
        self.filename = filename
        self.book = None
        self.encoding = kwargs.get('encoding') or 'utf8'
        try:
            self.book = xlrd.open_workbook(filename, encoding_override=self.encoding)
        except Exception:
            self.book = xlrd.open_workbook(filename, encoding_override='cp1251')
            self.sheet = self.book.sheet_by_index(0)

    def __formatrow__(self, types, values):
        #  Data Type Codes:
        #  EMPTY 0
        #  TEXT 1 a Unicode string
        #  NUMBER 2 float
        #  DATE 3 float
        #  BOOLEAN 4 int; 1 means TRUE, 0 means FALSE
        #  ERROR 5
        returnrow = []
        for i in range(len(types)):
            type, value = types[i], values[i]
            if type == 2:
                if value == int(value):
                    value = int(value)
            # elif type == 3:
            #    datetuple = xlrd.xldate_as_tuple(value, self.book.datemode)
            #    value = datetime.date(*datetuple[:3])
            elif type == 5:
                value = xlrd.error_text_from_code[value]
            returnrow.append(value)
        return returnrow

    @cached_property
    def count(self):
        if not self.book:
            return 0
        total = 0
        for sheet in self.book.sheets():
            total += sheet.nrows
        return total

    def __iter__(self):
        if not self.book:
            return
        for sheet in self.book.sheets():
            self.sheet = sheet
            for row in range(0, self.sheet.nrows):
                types, values = self.sheet.row_types(row), self.sheet.row_values(row)
                yield self.__formatrow__(types, values)

    def row(self, num):
        if not self.book:
            return []
        types, values = self.sheet.row_types(num), self.sheet.row_values(num)
        return self.__formatrow__(types, values)


class ExcelNewReader(object):

    def __init__(self, filename, **kwargs):
        if not os.path.isfile(filename):
            raise NameError("%s is not a valid filename" % filename)
        self.filename = filename
        self.book = load_workbook(filename, read_only=True, data_only=True)
        self.sheet = self.book.active
        self.encoding = 'utf8'

    def _correct(self, value):
        if value is None:
            return ''
        # fix problem when read integer(float) numbers
        if isinstance(value, float):
            value = R_0.sub('', str(value))
        return value

    @cached_property
    def count(self):
        if not self.book:
            return 0
        total = 0
        for sheet in self.book.worksheets:
            if sheet.max_row:
                total += sheet.max_row
        return total

    def __iter__(self):
        if not self.book:
            return
        for sheet in self.book.worksheets:
            self.sheet = sheet
            for row in self.sheet.iter_rows():
                result = [self._correct(cell.value) for cell in row]
                if result:
                    yield result

    def row(self, num):
        return [r.value for r in self.sheet.rows[num]]


class HTMLReader(object):

    def __init__(self, filename, **kwargs):
        self.filename = filename
        if not os.path.isfile(filename):
            raise NameError("%s is not a valid filename" % filename)
        self.filename = filename
        self.encoding = kwargs.get('encoding') or 'utf8'
        self.data = []
        self.__read()

    def __read(self):
        row = []
        file = open(self.filename, 'r')
        tr_begin = re.compile(r'\<tr.*?\>')
        tr_end = re.compile(r'\</tr\>')
        td = re.compile(r'\<td.*?\>(.+?)\</td\>')
        for line in file:
            res = tr_begin.search(line)
            if res:
                row = []
            res = td.search(line)
            if res:
                row.append(res.group(1))
            res = tr_end.search(line)
            if res and row:
                self.data.append(row)

    @cached_property
    def count(self):
        return len(self.data)

    def __iter__(self):
        for i in range(0, len(self.data)):
            yield self.data[i]

    def row(self, num):
        return self.data[num]


class CSVReader(object):

    def __init__(self, filename, **kwargs):
        self.filename = filename
        self.data = []
        with open(filename, 'rb') as f:
            head = b"".join([line for x, line in enumerate(f) if x < 500])
        self.encoding = detect_encoding(head)
        delimiter = kwargs.get('delimiter') or detect_delimiter(head)
        quoting = kwargs.get('quoting')
        if quoting is None:
            quoting = detect_quoting(head)
        else:
            quoting = int(quoting)
        if quoting == csv.QUOTE_NONE:
            quotechar = None
        else:
            quotechar = '"'
        delimiter = force_str(delimiter)
        f = open(filename, 'rU', encoding=self.encoding, errors='replace')
        self.csv_file = csv.reader(f,
                                   quoting=quoting,
                                   delimiter=delimiter,
                                   quotechar=quotechar)

    def __iter__(self):
        try:
            for row in self.csv_file:
                yield [force_str(r, encoding=self.encoding, errors="ignore") for r in row]
        except CSVError as e:
            if str(e) != 'newline inside string':
                raise e

    @cached_property
    def count(self):
        return count_lines(self.filename)

    def row(self, num):
        # not implemented
        return


class DBFReader(object):

    def __init__(self, filename, **kwargs):
        with open(filename, 'rb') as f:
            head = f.read(2048)
        self.encoding = kwargs.get('encoding') or detect_encoding(head, force_cp1251=True)
        self.table = dbf.Table(filename, codepage=self.encoding)
        self.table.use_deleted = False
        self.table.open()

    def __iter__(self):
        for row in self.table:
            result = []
            for r in row:
                if isinstance(r, str):
                    result.append(r.strip())
                else:
                    result.append(r)
            yield result

    @cached_property
    def count(self):
        if not self.table:
            return 0
        return len(self.table)


class Reader(object):

    def __init__(self, filename, **kwargs):
        ext = os.path.splitext(filename)[1].lower()
        self.reader = None
        self.errors = ''
        if not os.path.exists(filename):
            raise Exception('File "%s" does not exists.' % filename)
        if ext == '.xls':
            self.reader = ExcelReader(filename, **kwargs)
        elif ext == '.xlsx':
            self.reader = ExcelNewReader(filename)
        elif ext in ('.txt', '.csv', ''):
            self.reader = CSVReader(filename, **kwargs)
        elif ext == '.dbf':
            self.reader = DBFReader(filename, **kwargs)
        elif ext == '.html':
            self.reader = HTMLReader(filename, **kwargs)
        else:
            raise Exception('Reader: Wrong file "%s" format' % filename)

    def __iter__(self):
        return self.reader.__iter__()

    def parse(self, **kwargs):
        start = kwargs.get('start') or 0
        end = kwargs.get('end') or 0
        struct = kwargs.get('struct')
        required = kwargs.get('required', [])
        debug = kwargs.get('debug')
        blank = Struct(**dict((k, '') for k in struct.keys()))
        self.errors = []
        for index, row in enumerate(self.reader, start=1):
            ready = True
            result = copy.copy(blank)
            if start > index:
                continue
            if end and end < index:
                break
            for key, value in struct.items():
                if not value:
                    continue
                try:
                    result[key] = force_str(row[int(value) - 1], self.reader.encoding, errors='ignore').strip()
                    if not result[key] and key in required:
                        ready = False
                        if len(self.errors) < MAX_ERRORS:
                            self.errors.append(
                                _('Row %(index)s: empty field "%(key)s".') % {'index': index, 'key': key})
                        continue
                except IndexError:
                    if len(self.errors) < MAX_ERRORS:
                        self.errors.append(
                            _('Row %(index)s: check field "%(key)s".') % {'index': index, 'key': key})
            if debug:
                result._row = ';'.join(map(force_str, row))
                result._row_number = index
            if ready:
                yield result

    def row(self, num):
        if not self.reader:
            return
        return self.reader.row(num)

    @property
    def count(self):
        if not self.reader:
            return
        return self.reader.count

    def get_encoding(self):
        if not self.reader:
            return
        return self.reader.encoding

    def preview(self, max_row=30):
        column_count = 0
        items = []
        for i, row in enumerate(self):
            if i > max_row:
                break
            items.append(row)
            column_count = max(column_count, len(row))
        items.insert(0, range(1, column_count + 1))
        return items
